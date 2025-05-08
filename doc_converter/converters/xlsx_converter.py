"""
Excel-Konverter für die Dokumentenkonvertierungspipeline.
Konvertiert Excel-Tabellen (XLSX, XLS) in strukturierte Markdown-Dateien.
"""

import os
import re
from pathlib import Path
import logging
from typing import Dict, Any, List, Tuple, Optional, Set
import pandas as pd
import openpyxl
import numpy as np
from base64 import b64encode

from .base_converter import BaseConverter

class ExcelConverter(BaseConverter):
    """Konvertiert Excel-Tabellen (XLSX, XLS) in strukturierte Markdown-Dateien"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialisiert den Excel-Konverter.
        
        Args:
            config: Konfigurationswörterbuch mit Konvertereinstellungen
        """
        super().__init__(config)
        
        # Excel-spezifische Konfiguration
        self.extract_all_sheets = self.config.get('extract_all_sheets', True)
        self.extract_hidden_sheets = self.config.get('extract_hidden_sheets', False)
        self.max_rows = self.config.get('max_rows', 1000)  # Maximale Anzahl Zeilen pro Tabelle
        self.max_cols = self.config.get('max_cols', 20)    # Maximale Anzahl Spalten pro Tabelle
        self.include_sheet_names = self.config.get('include_sheet_names', True)
        self.convert_formulas = self.config.get('convert_formulas', True)
        self.extract_charts = self.config.get('extract_charts', False)  # Noch nicht implementiert
    
    def _convert_to_markdown(self, source_path: Path) -> Tuple[str, Dict[str, Any]]:
        """
        Konvertiert eine Excel-Datei in Markdown.
        
        Args:
            source_path: Pfad zur Excel-Datei
            
        Returns:
            Tuple aus (markdown_content, metadata)
        """
        self.logger.info(f"Starte Konvertierung von Excel: {source_path}")
        
        # Extrahiere Metadaten
        metadata = self._extract_excel_metadata(source_path)
        
        # Konvertiere Excel zu Markdown
        try:
            # Lade das Workbook
            wb = openpyxl.load_workbook(str(source_path), data_only=True, read_only=True)
            
            # Alternative mit Pandas (weniger Metadaten, aber robuster)
            # excel_data = pd.read_excel(source_path, sheet_name=None, engine='openpyxl')
            
            # Sammle Ergebnisse aller Tabellenblätter
            sheets_markdown = []
            sheet_names = wb.sheetnames
            
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                
                # Prüfe, ob das Blatt versteckt ist
                if sheet.sheet_state == 'hidden' and not self.extract_hidden_sheets:
                    self.logger.info(f"Überspringe verstecktes Blatt: {sheet_name}")
                    continue
                
                # Konvertiere das Tabellenblatt zu Markdown
                sheet_markdown = self._convert_sheet_to_markdown(sheet, sheet_name)
                
                if sheet_markdown:
                    sheets_markdown.append(f"## {sheet_name}\n\n{sheet_markdown}")
            
            # Kombiniere alle Tabellenblatt-Markdown-Texte
            markdown_content = "\n\n".join(sheets_markdown)
            
            self.logger.info(f"Excel-Konvertierung abgeschlossen: {len(markdown_content)} Zeichen")
            
            return markdown_content, metadata
            
        except Exception as e:
            self.logger.error(f"Fehler bei der Excel-Konvertierung: {e}", exc_info=True)
            
            # Fallback mit Pandas
            try:
                self.logger.info("Versuche Fallback mit Pandas")
                return self._convert_with_pandas(source_path, metadata)
            except Exception as fallback_error:
                self.logger.error(f"Auch Fallback mit Pandas fehlgeschlagen: {fallback_error}", exc_info=True)
                return f"# Konvertierung fehlgeschlagen\n\nFehler: {str(e)}\nFallback-Fehler: {str(fallback_error)}", metadata
    
    def _convert_with_pandas(self, source_path: Path, metadata: Dict[str, Any]) -> Tuple[str, Dict[str, Any]]:
        """
        Konvertiert eine Excel-Datei mit Pandas (Fallback-Methode).
        
        Args:
            source_path: Pfad zur Excel-Datei
            metadata: Bereits extrahierte Metadaten
            
        Returns:
            Tuple aus (markdown_content, metadata)
        """
        try:
            # Lade alle Tabellenblätter
            excel_data = pd.read_excel(source_path, sheet_name=None)
            
            sheets_markdown = []
            
            for sheet_name, df in excel_data.items():
                # Begrenze Größe
                df_limited = df.iloc[:min(len(df), self.max_rows), :min(df.shape[1], self.max_cols)]
                
                # Tabelle zu Markdown konvertieren
                table_md = self._dataframe_to_markdown(df_limited)
                
                if table_md:
                    # Füge Tabellenblattname hinzu
                    if self.include_sheet_names:
                        sheet_md = f"## {sheet_name}\n\n{table_md}"
                    else:
                        sheet_md = table_md
                    
                    sheets_markdown.append(sheet_md)
            
            # Kombiniere alle Tabellenblatt-Markdown-Texte
            markdown_content = "\n\n".join(sheets_markdown)
            
            if not markdown_content:
                return "# Leere Excel-Datei\n\nDie Excel-Datei enthält keine konvertierbaren Daten.", metadata
            
            return markdown_content, metadata
            
        except Exception as e:
            self.logger.error(f"Fehler bei der Pandas-Konvertierung: {e}", exc_info=True)
            return f"# Konvertierung fehlgeschlagen\n\nFehler: {str(e)}", metadata
    
    def _convert_sheet_to_markdown(self, sheet, sheet_name: str) -> str:
        """
        Konvertiert ein Excel-Tabellenblatt in Markdown.
        
        Args:
            sheet: openpyxl Worksheet-Objekt
            sheet_name: Name des Tabellenblatts
            
        Returns:
            Markdown-Repräsentation des Tabellenblatts
        """
        try:
            # Bestimme den benutzten Bereich
            used_range = self._get_used_range(sheet)
            
            if not used_range:
                return ""
            
            min_row, max_row, min_col, max_col = used_range
            
            # Begrenze die Größe
            max_row = min(max_row, min_row + self.max_rows - 1)
            max_col = min(max_col, min_col + self.max_cols - 1)
            
            # Sammle Zelldaten
            rows = []
            for row_idx in range(min_row, max_row + 1):
                row_data = []
                for col_idx in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    # Konvertiere None zu leerem String
                    if value is None:
                        value = ""
                    # Konvertiere Zahlen
                    elif isinstance(value, (int, float)):
                        if value == int(value):  # Ganze Zahl
                            value = str(int(value))
                        else:
                            value = str(value)
                    # Konvertiere alles andere zu String
                    else:
                        value = str(value)
                    
                    row_data.append(value)
                
                rows.append(row_data)
            
            # Erstelle Markdown-Tabelle
            if not rows:
                return ""
            
            table_rows = []
            
            # Header-Zeile
            header_row = "| " + " | ".join(rows[0]) + " |"
            table_rows.append(header_row)
            
            # Trennzeile
            separator_row = "| " + " | ".join(["---"] * len(rows[0])) + " |"
            table_rows.append(separator_row)
            
            # Datenzeilen
            for row in rows[1:]:
                data_row = "| " + " | ".join(row) + " |"
                table_rows.append(data_row)
            
            return "\n".join(table_rows)
            
        except Exception as e:
            self.logger.error(f"Fehler bei der Konvertierung des Tabellenblatts {sheet_name}: {e}", exc_info=True)
            return f"*Fehler bei der Konvertierung des Tabellenblatts {sheet_name}: {str(e)}*"
    
    def _get_used_range(self, sheet) -> Optional[Tuple[int, int, int, int]]:
        """
        Bestimmt den benutzten Bereich in einem Tabellenblatt.
        
        Args:
            sheet: openpyxl Worksheet-Objekt
            
        Returns:
            Tuple aus (min_row, max_row, min_col, max_col) oder None, wenn das Blatt leer ist
        """
        try:
            if hasattr(sheet, 'calculate_dimension') and callable(sheet.calculate_dimension):
                # Für normale Worksheets
                dimension = sheet.calculate_dimension()
                if dimension == 'A1:A1' and sheet['A1'].value is None:
                    # Leeres Blatt
                    return None
                
                # Parse Dimension (z.B. 'A1:C10')
                match = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', dimension)
                if match:
                    min_col_letter, min_row, max_col_letter, max_row = match.groups()
                    min_row = int(min_row)
                    max_row = int(max_row)
                    
                    # Konvertiere Spaltenbuchstaben zu Zahlen
                    min_col = self._column_letter_to_number(min_col_letter)
                    max_col = self._column_letter_to_number(max_col_letter)
                    
                    return min_row, max_row, min_col, max_col
            
            # Fallback für read_only Worksheets oder wenn calculate_dimension fehlschlägt
            min_row = min(sheet.min_row, 1)
            max_row = sheet.max_row
            min_col = min(sheet.min_column, 1)
            max_col = sheet.max_column
            
            if min_row > max_row or min_col > max_col:
                return None
            
            return min_row, max_row, min_col, max_col
            
        except Exception as e:
            self.logger.warning(f"Fehler bei der Bestimmung des benutzten Bereichs: {e}")
            return 1, 50, 1, 10  # Standardbereich als Fallback
    
    def _column_letter_to_number(self, column_letter: str) -> int:
        """
        Konvertiert einen Spaltenbuchstaben (z.B. 'A', 'AB') in eine Zahl.
        
        Args:
            column_letter: Spaltenbuchstabe
            
        Returns:
            Spaltennummer (1-basiert)
        """
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def _dataframe_to_markdown(self, df: pd.DataFrame) -> str:
        """
        Konvertiert einen Pandas DataFrame in eine Markdown-Tabelle.
        
        Args:
            df: Pandas DataFrame
            
        Returns:
            Markdown-Tabelle
        """
        # Überprüfe, ob der DataFrame leer ist
        if df.empty:
            return ""
        
        # Erstelle Header-Zeile
        header = "| " + " | ".join(str(col) for col in df.columns) + " |"
        
        # Erstelle Trennzeile
        separator = "| " + " | ".join(["---"] * len(df.columns)) + " |"
        
        # Erstelle Datenzeilen
        rows = []
        for _, row in df.iterrows():
            values = []
            for val in row:
                if pd.isna(val):
                    values.append("")
                elif isinstance(val, (int, float)) and val == int(val):
                    values.append(str(int(val)))
                else:
                    # Ersetze Zeilenumbrüche durch Leerzeichen
                    val_str = str(val).replace('\n', ' ').replace('\r', '')
                    values.append(val_str)
            
            rows.append("| " + " | ".join(values) + " |")
        
        # Kombiniere alle Teile
        return "\n".join([header, separator] + rows)
    
    def _extract_excel_metadata(self, source_path: Path) -> Dict[str, Any]:
        """
        Extrahiert Metadaten aus einer Excel-Datei.
        
        Args:
            source_path: Pfad zur Excel-Datei
            
        Returns:
            Dictionary mit Metadaten
        """
        metadata = {
            'title': source_path.stem,
            'original_format': 'XLSX' if source_path.suffix.lower() == '.xlsx' else 'XLS',
            'has_tables': True,
            'has_images': False,  # Standardwert, wird unten aktualisiert
            'has_charts': False,  # Standardwert, wird unten aktualisiert
            'sheets': []
        }
        
        try:
            # Lade Workbook für Metadaten
            wb = openpyxl.load_workbook(
                str(source_path), 
                read_only=True, 
                keep_vba=False, 
                data_only=True,
                keep_links=False
            )
            
            # Dokumenteigenschaften
            props = wb.properties
            if props:
                if props.title:
                    metadata['title'] = props.title
                if props.creator:
                    metadata['author'] = props.creator
                if props.created:
                    metadata['created'] = str(props.created)
                if props.modified:
                    metadata['modified'] = str(props.modified)
                if props.subject:
                    metadata['subject'] = props.subject
                if props.keywords:
                    metadata['keywords'] = props.keywords
            
            # Tabellenblätter
            sheet_names = wb.sheetnames
            metadata['sheet_count'] = len(sheet_names)
            
            sheet_info = []
            for sheet_name in sheet_names:
                sheet = wb[sheet_name]
                
                # Basisinformationen über das Tabellenblatt
                sheet_data = {
                    'name': sheet_name,
                    'hidden': sheet.sheet_state == 'hidden',
                    'has_data': False,
                    'row_count': 0,
                    'column_count': 0
                }
                
                # Bestimme Bereichsgröße
                used_range = self._get_used_range(sheet)
                if used_range:
                    min_row, max_row, min_col, max_col = used_range
                    sheet_data['has_data'] = True
                    sheet_data['row_count'] = max_row - min_row + 1
                    sheet_data['column_count'] = max_col - min_col + 1
                
                # Prüfe auf Bilder
                if hasattr(sheet, '_images'):
                    sheet_data['has_images'] = len(sheet._images) > 0
                    if sheet_data['has_images']:
                        metadata['has_images'] = True
                
                # Prüfe auf Diagramme
                if hasattr(sheet, '_charts'):
                    sheet_data['has_charts'] = len(sheet._charts) > 0
                    if sheet_data['has_charts']:
                        metadata['has_charts'] = True
                
                sheet_info.append(sheet_data)
            
            metadata['sheets'] = sheet_info
            
            # Berechne Gesamtzahl der Zeilen und Spalten
            data_sheets = [s for s in sheet_info if s['has_data']]
            if data_sheets:
                metadata['total_rows'] = sum(s['row_count'] for s in data_sheets)
                metadata['total_columns'] = sum(s['column_count'] for s in data_sheets)
            else:
                metadata['total_rows'] = 0
                metadata['total_columns'] = 0
            
            # Schätze Seitenzahl (grobe Annäherung)
            metadata['pages'] = max(1, metadata['total_rows'] // 40)
            
        except Exception as e:
            self.logger.warning(f"Fehler beim Extrahieren der Excel-Metadaten: {e}")
            
            try:
                # Fallback mit Pandas
                excel_data = pd.read_excel(source_path, sheet_name=None)
                metadata['sheet_count'] = len(excel_data)
                metadata['sheets'] = [{'name': name, 'row_count': df.shape[0], 'column_count': df.shape[1]} 
                                     for name, df in excel_data.items()]
                metadata['total_rows'] = sum(df.shape[0] for df in excel_data.values())
                metadata['total_columns'] = sum(df.shape[1] for df in excel_data.values())
                metadata['pages'] = max(1, metadata['total_rows'] // 40)
            except Exception as fallback_error:
                self.logger.error(f"Auch Fallback für Metadaten fehlgeschlagen: {fallback_error}")
        
        return metadata


if __name__ == "__main__":
    # Setup Logging
    logging.basicConfig(level=logging.INFO, 
                        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    # Test des Excel-Konverters
    import sys
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
        target_dir = sys.argv[2] if len(sys.argv) > 2 else "output"
        
        converter = ExcelConverter()
        result = converter.convert(excel_path, target_dir)
        
        if result['success']:
            print(f"Konvertierung erfolgreich: {result['target']}")
        else:
            print(f"Fehler bei der Konvertierung: {result.get('error', 'Unbekannter Fehler')}")
    else:
        print("Verwendung: python xlsx_converter.py <excel_datei> [ziel_verzeichnis]")